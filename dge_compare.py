#!/usr/bin/env python3
"""Compare DGE XLSX reports produced by Mongo/SFTP and Fabric.

The comparator intentionally checks workbook values, not XLSX bytes. XLSX files
can differ in styles, metadata, compression, and calculation state while still
holding the same report data.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import re
from collections import Counter
from dataclasses import dataclass
from datetime import UTC, date, datetime, time, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


REPORT_ALIASES = {
    "game_round_report": ("game_round_report",),
    "game_summary_report": ("game_summary_report",),
    "machine_summary_report": ("machine_summary_report",),
    "pending_transaction": ("pending_transaction",),
    "session_transaction": ("session_transaction",),
    "void_transaction": ("void_transaction", "void-transaction"),
}

REPORTS = tuple(REPORT_ALIASES.keys())
DATE_RE = re.compile(r"(\d{8})(?!\d)")
DATE_RANGE_PREFIX = "date range:"
DEFAULT_DIFF_LIMIT = 50
SIMILAR_ROW_THRESHOLD = 0.6
METADATA_LABELS = {
    "title": "Top header row 1 / report title",
    "date_range": "Top header row 2 / date range",
    "sheets": "Worksheet name",
    "headers": "Column header row",
}
IDENTITY_HEADERS = (
    "Transaction ID",
    "Round ID",
    "Session ID",
    "Patron ID",
    "Machine ID",
    "Machine Name",
    "Game ID",
    "Game Name",
    "Gaming Date",
    "Operator name",
)
WEAK_IDENTITY_HEADERS = {"Gaming Date", "Operator name"}


@dataclass(frozen=True)
class ReportKey:
    report: str
    brand: str
    report_date: date

    def as_id(self) -> str:
        return f"{self.report}|{self.brand}|{self.report_date:%Y%m%d}"


@dataclass(frozen=True)
class DiscoveredFile:
    key: ReportKey
    path: Path
    modified_at: datetime


@dataclass
class WorkbookData:
    title: str
    date_range: str
    sheet_names: list[str]
    headers: list[str]
    rows: list[list[str]]


@dataclass(frozen=True)
class RunContext:
    run_id: str
    generated_at: str
    file_suffix: str
    business_date_range: str
    brand_label: str
    report_label: str
    mode: str
    row_order_policy: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compare DGE XLSX reports from a Mongo/SFTP folder and a Fabric folder.",
    )
    parser.add_argument("--sftp-dir", required=True, type=Path, help="Folder with Mongo/SFTP-produced XLSX reports.")
    parser.add_argument("--fabric-dir", required=True, type=Path, help="Folder with Fabric-produced XLSX reports.")
    parser.add_argument("--output-dir", required=True, type=Path, help="Folder for summary.csv, details.csv, JSON, and HTML dashboard.")
    parser.add_argument("--date", action="append", dest="dates", help="Business/report date to compare, YYYYMMDD. Repeatable.")
    parser.add_argument("--from-date", help="First business/report date to compare, YYYYMMDD. Inclusive.")
    parser.add_argument("--until-date", help="Last business/report date to compare, YYYYMMDD. Inclusive.")
    parser.add_argument("--days", type=int, default=2, help="Number of days to compare when --date is not provided. Default: 2.")
    parser.add_argument("--end-date", help="Last business date for --days, YYYYMMDD. Default: yesterday.")
    parser.add_argument("--brand", action="append", dest="brands", help="Expected brand, for example caesars-nj. Repeatable.")
    parser.add_argument("--brand-alias", action="append", dest="brand_aliases", help="Map a file brand to the comparison brand, source=target. Repeatable, for example fanduel-nj=fd.")
    parser.add_argument("--report", action="append", dest="reports", choices=REPORTS, help="Report to compare. Repeatable. Default: all.")
    parser.add_argument("--sftp-source-url", help="Optional human link to the SFTP/source folder for the dashboard.")
    parser.add_argument("--fabric-source-url", help="Optional human link to the Fabric/source folder for the dashboard.")
    parser.add_argument("--created-from", default=None, help="Optional local file modified-time lower bound, HH:MM.")
    parser.add_argument("--created-to", default=None, help="Optional local file modified-time upper bound, HH:MM.")
    parser.add_argument("--strict-order", action="store_true", help="Require row order to match exactly.")
    parser.add_argument("--diff-limit", type=int, default=DEFAULT_DIFF_LIMIT, help="Maximum row-level differences to write per file pair.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    target_dates = parse_target_dates(args)
    modified_window = parse_time_window(args.created_from, args.created_to)

    brand_aliases = parse_brand_aliases(args.brand_aliases or [])
    sftp_files = normalize_file_brands(discover_files(args.sftp_dir, modified_window), brand_aliases)
    fabric_files = normalize_file_brands(discover_files(args.fabric_dir, modified_window), brand_aliases)
    brands = sorted({normalize_brand(brand, brand_aliases) for brand in args.brands} if args.brands else {f.key.brand for f in [*sftp_files, *fabric_files]})
    reports = tuple(args.reports or REPORTS)

    if not brands:
        raise SystemExit("No brands found. Pass --brand or place matching XLSX files in the input folders.")

    sftp_index = choose_latest(sftp_files)
    fabric_index = choose_latest(fabric_files)
    expected_keys = [
        ReportKey(report=report, brand=brand, report_date=report_date)
        for report_date in target_dates
        for brand in brands
        for report in reports
    ]

    summary_rows: list[dict[str, object]] = []
    detail_rows: list[dict[str, object]] = []
    run_context = build_run_context(
        target_dates=target_dates,
        brands=brands,
        reports=reports,
        strict_order=args.strict_order,
    )

    for key in expected_keys:
        result, details = compare_pair(
            key=key,
            sftp_file=sftp_index.get(key),
            fabric_file=fabric_index.get(key),
            strict_order=args.strict_order,
            diff_limit=args.diff_limit,
        )
        summary_rows.append(result)
        detail_rows.extend(details)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    artifact_paths = dashboard_artifact_paths(args.output_dir, run_context.file_suffix)

    write_csv(artifact_paths["summary_latest"], summary_rows)
    write_csv(artifact_paths["summary_archived"], summary_rows)
    write_csv(artifact_paths["details_latest"], detail_rows)
    write_csv(artifact_paths["details_archived"], detail_rows)
    write_json(artifact_paths["summary_json_latest"], summary_rows)
    write_json(artifact_paths["summary_json_archived"], summary_rows)
    write_json(artifact_paths["details_json_latest"], detail_rows)
    write_json(artifact_paths["details_json_archived"], detail_rows)

    write_dashboard(
        artifact_paths["dashboard_latest"],
        summary_rows,
        detail_rows,
        run_context=run_context,
        artifact_paths=artifact_paths,
        sftp_source_url=args.sftp_source_url,
        fabric_source_url=args.fabric_source_url,
    )
    write_dashboard(
        artifact_paths["dashboard_archived"],
        summary_rows,
        detail_rows,
        run_context=run_context,
        artifact_paths=artifact_paths,
        sftp_source_url=args.sftp_source_url,
        fabric_source_url=args.fabric_source_url,
    )

    status_counts = Counter(str(row["status"]) for row in summary_rows)
    print(f"Wrote latest dashboard: {artifact_paths['dashboard_latest']}")
    print(f"Wrote run dashboard: {artifact_paths['dashboard_archived']}")
    print("Status counts: " + ", ".join(f"{status}={count}" for status, count in sorted(status_counts.items())))
    return 0 if not any(status not in {"MATCH"} for status in status_counts) else 2


def parse_target_dates(args: argparse.Namespace) -> list[date]:
    if args.dates:
        return [datetime.strptime(value, "%Y%m%d").date() for value in args.dates]

    if args.from_date or args.until_date:
        if not args.from_date or not args.until_date:
            raise SystemExit("Pass both --from-date and --until-date, or neither.")
        from_date = datetime.strptime(args.from_date, "%Y%m%d").date()
        until_date = datetime.strptime(args.until_date, "%Y%m%d").date()
        if from_date > until_date:
            raise SystemExit("--from-date must be before or equal to --until-date.")
        days = (until_date - from_date).days
        return [from_date + timedelta(days=offset) for offset in range(days + 1)]

    if args.end_date:
        end_date = datetime.strptime(args.end_date, "%Y%m%d").date()
    else:
        end_date = date.today() - timedelta(days=1)

    return [end_date - timedelta(days=offset) for offset in range(args.days)]


def parse_brand_aliases(values: Iterable[str]) -> dict[str, str]:
    aliases: dict[str, str] = {}
    for value in values:
        if "=" not in value:
            raise SystemExit(f"Invalid --brand-alias '{value}'. Use source=target.")
        source, target = (part.strip() for part in value.split("=", 1))
        if not source or not target:
            raise SystemExit(f"Invalid --brand-alias '{value}'. Use source=target.")
        aliases[source.lower()] = target
    return aliases


def normalize_brand(brand: str, aliases: dict[str, str]) -> str:
    return aliases.get(brand.lower(), brand)


def normalize_file_brands(files: list[DiscoveredFile], aliases: dict[str, str]) -> list[DiscoveredFile]:
    if not aliases:
        return files

    normalized: list[DiscoveredFile] = []
    for file in files:
        brand = normalize_brand(file.key.brand, aliases)
        if brand == file.key.brand:
            normalized.append(file)
            continue
        normalized.append(
            DiscoveredFile(
                key=ReportKey(
                    report=file.key.report,
                    brand=brand,
                    report_date=file.key.report_date,
                ),
                path=file.path,
                modified_at=file.modified_at,
            )
        )
    return normalized


def build_run_context(
    target_dates: list[date],
    brands: list[str],
    reports: Iterable[str],
    strict_order: bool,
) -> RunContext:
    generated_at = utc_now_iso()
    run_started = datetime.now(UTC)
    run_id = run_started.strftime("%Y%m%d_%H%M%S")

    first_date = min(target_dates)
    last_date = max(target_dates)
    if first_date == last_date:
        business_date_range = f"{first_date:%Y-%m-%d}"
        date_slug = f"{first_date:%Y%m%d}"
    else:
        business_date_range = f"{first_date:%Y-%m-%d} to {last_date:%Y-%m-%d}"
        date_slug = f"{first_date:%Y%m%d}-{last_date:%Y%m%d}"

    brand_label = ", ".join(brands) if len(brands) <= 3 else f"{len(brands)} brands"
    brand_slug = sanitize_filename_part(brands[0]) if len(brands) == 1 else f"{len(brands)}-brands"
    report_list = tuple(reports)
    report_label = ", ".join(report_list) if len(report_list) <= 3 else f"{len(report_list)} reports"
    mode = "Daily monitoring" if len(target_dates) == 1 else "Investigation"
    row_order_policy = "Strict row order comparison" if strict_order else "Row order ignored"

    return RunContext(
        run_id=run_id,
        generated_at=generated_at,
        file_suffix=f"{brand_slug}_{date_slug}_{run_id}",
        business_date_range=business_date_range,
        brand_label=brand_label,
        report_label=report_label,
        mode=mode,
        row_order_policy=row_order_policy,
    )


def sanitize_filename_part(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "-", value.strip())
    return cleaned.strip("-") or "run"


def dashboard_artifact_paths(output_dir: Path, suffix: str) -> dict[str, Path]:
    return {
        "dashboard_latest": output_dir / "dashboard.html",
        "dashboard_archived": output_dir / f"dashboard_{suffix}.html",
        "summary_latest": output_dir / "summary.csv",
        "summary_archived": output_dir / f"summary_{suffix}.csv",
        "details_latest": output_dir / "details.csv",
        "details_archived": output_dir / f"details_{suffix}.csv",
        "summary_json_latest": output_dir / "summary.json",
        "summary_json_archived": output_dir / f"summary_{suffix}.json",
        "details_json_latest": output_dir / "details.json",
        "details_json_archived": output_dir / f"details_{suffix}.json",
    }


def parse_time_window(created_from: str | None, created_to: str | None) -> tuple[time, time] | None:
    if not created_from and not created_to:
        return None
    if not created_from or not created_to:
        raise SystemExit("Pass both --created-from and --created-to, or neither.")
    return (datetime.strptime(created_from, "%H:%M").time(), datetime.strptime(created_to, "%H:%M").time())


def discover_files(root: Path, modified_window: tuple[time, time] | None) -> list[DiscoveredFile]:
    if not root.exists():
        return []

    files: list[DiscoveredFile] = []
    for path in root.rglob("*.xlsx"):
        modified_at = datetime.fromtimestamp(path.stat().st_mtime)
        if modified_window and not is_time_in_window(modified_at.time(), modified_window):
            continue
        key = parse_report_key(path)
        if key is None:
            continue
        files.append(DiscoveredFile(key=key, path=path, modified_at=modified_at))
    return files


def is_time_in_window(value: time, window: tuple[time, time]) -> bool:
    start, end = window
    if start <= end:
        return start <= value <= end
    return value >= start or value <= end


def parse_report_key(path: Path) -> ReportKey | None:
    stem = path.stem
    date_matches = list(DATE_RE.finditer(stem))
    if not date_matches:
        return None

    date_match = date_matches[-1]
    report_date = datetime.strptime(date_match.group(1), "%Y%m%d").date()
    prefix = stem[: date_match.start()].strip("_-. ")

    for canonical, aliases in REPORT_ALIASES.items():
        for alias in aliases:
            pattern = re.compile("^" + r"[-_]+".join(map(re.escape, alias.split("_"))), re.IGNORECASE)
            match = pattern.match(prefix)
            if match:
                brand = prefix[match.end() :].strip("_-. ")
                if brand:
                    return ReportKey(report=canonical, brand=brand, report_date=report_date)
    return None


def choose_latest(files: Iterable[DiscoveredFile]) -> dict[ReportKey, DiscoveredFile]:
    latest: dict[ReportKey, DiscoveredFile] = {}
    for file in files:
        current = latest.get(file.key)
        if current is None or file.modified_at > current.modified_at:
            latest[file.key] = file
    return latest


def compare_pair(
    key: ReportKey,
    sftp_file: DiscoveredFile | None,
    fabric_file: DiscoveredFile | None,
    strict_order: bool,
    diff_limit: int,
) -> tuple[dict[str, object], list[dict[str, object]]]:
    result = {
        "report_date": f"{key.report_date:%Y-%m-%d}",
        "brand": key.brand,
        "report": key.report,
        "status": "MATCH",
        "sftp_path": str(sftp_file.path) if sftp_file else "",
        "fabric_path": str(fabric_file.path) if fabric_file else "",
        "sftp_rows": "",
        "fabric_rows": "",
        "sftp_columns": "",
        "fabric_columns": "",
        "difference_count": 0,
        "comments": "",
        "checked_at": utc_now_iso(),
    }
    details: list[dict[str, object]] = []

    if sftp_file is None or fabric_file is None:
        status = "MISSING_IN_SFTP" if sftp_file is None and fabric_file is not None else "MISSING_IN_FABRIC"
        if sftp_file is None and fabric_file is None:
            status = "MISSING_BOTH"
        result["status"] = status
        result["difference_count"] = 1
        details.append(detail_row(key, "file", status, "", "", ""))
        result["comments"] = summarize_details(details)
        return result, details

    sftp_data = read_workbook(sftp_file.path)
    fabric_data = read_workbook(fabric_file.path)
    result["sftp_rows"] = len(sftp_data.rows)
    result["fabric_rows"] = len(fabric_data.rows)
    result["sftp_columns"] = len(sftp_data.headers)
    result["fabric_columns"] = len(fabric_data.headers)

    add_metadata_diff(details, key, "title", sftp_data.title, fabric_data.title)
    add_metadata_diff(details, key, "date_range", sftp_data.date_range, fabric_data.date_range)
    add_metadata_diff(details, key, "sheets", " | ".join(sftp_data.sheet_names), " | ".join(fabric_data.sheet_names))
    add_metadata_diff(details, key, "headers", json.dumps(sftp_data.headers), json.dumps(fabric_data.headers))

    if strict_order:
        compare_ordered_rows(details, key, sftp_data.headers, sftp_data.rows, fabric_data.rows, diff_limit)
    else:
        compare_row_multisets(details, key, sftp_data.headers, sftp_data.rows, fabric_data.rows, diff_limit)

    result["difference_count"] = len(details)
    result["status"] = "MATCH" if not details else "MISMATCH"
    result["comments"] = summarize_details(details)
    return result, details


def summarize_details(details: list[dict[str, object]]) -> str:
    if not details:
        return ""

    counts = Counter(str(detail["diff_type"]) for detail in details)
    comments: list[str] = []
    if counts.get("file"):
        comments.append("file missing")
    if counts.get("metadata"):
        comments.append(f"{counts['metadata']} workbook/header difference(s)")
    if counts.get("row_count"):
        comments.append("row count differs")
    if counts.get("cell"):
        comments.append(f"{counts['cell']} changed field(s)")
    row_diffs = counts.get("row", 0)
    if row_diffs:
        comments.append(f"{row_diffs} unmatched row(s)")
    return "; ".join(comments)


def read_workbook(path: Path) -> WorkbookData:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_names = list(workbook.sheetnames)
    sheet = workbook[sheet_names[0]]
    rows = list(sheet.iter_rows(values_only=False))
    header_idx = find_header_row(rows)

    title = normalize_cell(rows[0][0]) if rows and rows[0] else ""
    date_range = normalize_cell(rows[1][0]) if len(rows) > 1 and rows[1] else ""
    headers = [normalize_cell(cell) for cell in rows[header_idx]]
    data_rows = [
        [normalize_cell(cell) for cell in row[: len(headers)]]
        for row in rows[header_idx + 1 :]
        if any(normalize_cell(cell) != "" for cell in row)
    ]

    workbook.close()
    return WorkbookData(title=title, date_range=date_range, sheet_names=sheet_names, headers=headers, rows=data_rows)


def find_header_row(rows: list[tuple]) -> int:
    for index, row in enumerate(rows):
        values = [normalize_cell(cell) for cell in row]
        non_empty = [value for value in values if value != ""]
        if len(non_empty) >= 2 and not values[0].lower().startswith(DATE_RANGE_PREFIX):
            if index > 0 and len(non_empty) > 1:
                return index
    return 0


def normalize_cell(cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, datetime):
        number_format = (cell.number_format or "").lower()
        if any(token in number_format for token in ("h", "s")):
            return value.isoformat(timespec="milliseconds")
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return normalize_number(value)
    if isinstance(value, Decimal):
        return normalize_decimal(value)
    return str(value)


def normalize_number(value: float) -> str:
    if value.is_integer():
        return str(int(value))
    return normalize_decimal(Decimal(str(value)))


def normalize_decimal(value: Decimal) -> str:
    normalized = value.normalize()
    if normalized == normalized.to_integral():
        return str(normalized.to_integral())
    return format(normalized, "f")


def add_metadata_diff(details: list[dict[str, object]], key: ReportKey, field: str, sftp_value: str, fabric_value: str) -> None:
    if sftp_value != fabric_value:
        details.append(detail_row(key, "metadata", METADATA_LABELS.get(field, field), sftp_value, fabric_value, ""))


def compare_ordered_rows(
    details: list[dict[str, object]],
    key: ReportKey,
    headers: list[str],
    sftp_rows: list[list[str]],
    fabric_rows: list[list[str]],
    diff_limit: int,
) -> None:
    if len(sftp_rows) != len(fabric_rows):
        details.append(detail_row(key, "row_count", "row_count", str(len(sftp_rows)), str(len(fabric_rows)), ""))

    for index in range(max(len(sftp_rows), len(fabric_rows))):
        if len(details) >= diff_limit:
            return
        sftp_row = sftp_rows[index] if index < len(sftp_rows) else None
        fabric_row = fabric_rows[index] if index < len(fabric_rows) else None
        if sftp_row is None:
            details.append(detail_row(
                key, "row", "extra_in_fabric", "", json.dumps(fabric_row, ensure_ascii=False), f"row={index + 1}",
                row_key=row_label(headers, fabric_row),
            ))
        elif fabric_row is None:
            details.append(detail_row(
                key, "row", "missing_from_fabric", json.dumps(sftp_row, ensure_ascii=False), "", f"row={index + 1}",
                row_key=row_label(headers, sftp_row),
            ))
        elif sftp_row != fabric_row:
            append_cell_diffs(details, key, headers, sftp_row, fabric_row, diff_limit, row_group=f"ordered-row-{index + 1}")


def compare_row_multisets(
    details: list[dict[str, object]],
    key: ReportKey,
    headers: list[str],
    sftp_rows: list[list[str]],
    fabric_rows: list[list[str]],
    diff_limit: int,
) -> None:
    if len(sftp_rows) != len(fabric_rows):
        details.append(detail_row(key, "row_count", "row_count", str(len(sftp_rows)), str(len(fabric_rows)), ""))

    sftp_counter = Counter(tuple(row) for row in sftp_rows)
    fabric_counter = Counter(tuple(row) for row in fabric_rows)

    missing_from_fabric = sftp_counter - fabric_counter
    extra_in_fabric = fabric_counter - sftp_counter
    missing_rows = expand_counter(missing_from_fabric)
    extra_rows = expand_counter(extra_in_fabric)
    pairings, unmatched_sftp, unmatched_fabric = pair_similar_rows(missing_rows, extra_rows)

    for pair_index, (sftp_row, fabric_row) in enumerate(pairings, start=1):
        if len(details) >= diff_limit:
            return
        append_cell_diffs(details, key, headers, sftp_row, fabric_row, diff_limit, row_group=f"paired-row-{pair_index}")

    for row in unmatched_sftp:
        if len(details) >= diff_limit:
            return
        details.append(detail_row(
            key,
            "row",
            "missing_from_fabric",
            json.dumps(row, ensure_ascii=False),
            "",
            "no matching Fabric row found",
            row_key=row_label(headers, row),
        ))

    for row in unmatched_fabric:
        if len(details) >= diff_limit:
            return
        details.append(detail_row(
            key,
            "row",
            "extra_in_fabric",
            "",
            json.dumps(row, ensure_ascii=False),
            "no matching SFTP row found",
            row_key=row_label(headers, row),
        ))


def expand_counter(counter: Counter[tuple[str, ...]]) -> list[list[str]]:
    rows: list[list[str]] = []
    for row, count in counter.items():
        rows.extend([list(row) for _ in range(count)])
    return rows


def pair_similar_rows(sftp_rows: list[list[str]], fabric_rows: list[list[str]]) -> tuple[list[tuple[list[str], list[str]]], list[list[str]], list[list[str]]]:
    candidates: list[tuple[float, int, int]] = []
    for sftp_index, sftp_row in enumerate(sftp_rows):
        for fabric_index, fabric_row in enumerate(fabric_rows):
            score = row_similarity(sftp_row, fabric_row)
            if score >= SIMILAR_ROW_THRESHOLD:
                candidates.append((score, sftp_index, fabric_index))

    paired_sftp: set[int] = set()
    paired_fabric: set[int] = set()
    pair_records: list[tuple[int, int, list[str], list[str]]] = []
    for _, sftp_index, fabric_index in sorted(candidates, reverse=True):
        if sftp_index in paired_sftp or fabric_index in paired_fabric:
            continue
        paired_sftp.add(sftp_index)
        paired_fabric.add(fabric_index)
        pair_records.append((sftp_index, fabric_index, sftp_rows[sftp_index], fabric_rows[fabric_index]))

    pairings = [(sftp_row, fabric_row) for _, _, sftp_row, fabric_row in sorted(pair_records)]
    unmatched_sftp = [row for index, row in enumerate(sftp_rows) if index not in paired_sftp]
    unmatched_fabric = [row for index, row in enumerate(fabric_rows) if index not in paired_fabric]
    return pairings, unmatched_sftp, unmatched_fabric


def row_similarity(sftp_row: list[str], fabric_row: list[str]) -> float:
    width = max(len(sftp_row), len(fabric_row))
    if width == 0:
        return 1.0
    matches = 0
    for index in range(width):
        sftp_value = sftp_row[index] if index < len(sftp_row) else ""
        fabric_value = fabric_row[index] if index < len(fabric_row) else ""
        if sftp_value == fabric_value:
            matches += 1
    return matches / width


def append_cell_diffs(
    details: list[dict[str, object]],
    key: ReportKey,
    headers: list[str],
    sftp_row: list[str],
    fabric_row: list[str],
    diff_limit: int,
    row_group: str,
) -> None:
    row_key = paired_row_label(headers, sftp_row, fabric_row)
    width = max(len(sftp_row), len(fabric_row), len(headers))
    for index in range(width):
        if len(details) >= diff_limit:
            return
        sftp_value = sftp_row[index] if index < len(sftp_row) else ""
        fabric_value = fabric_row[index] if index < len(fabric_row) else ""
        if sftp_value == fabric_value:
            continue
        field = headers[index] if index < len(headers) and headers[index] else f"Column {index + 1}"
        details.append(detail_row(
            key,
            "cell",
            field,
            sftp_value,
            fabric_value,
            "",
            row_key=row_key,
            row_group=row_group,
        ))


def paired_row_label(headers: list[str], sftp_row: list[str], fabric_row: list[str]) -> str:
    shared_parts: list[str] = []
    strong_shared_parts: list[str] = []
    changed_parts: list[str] = []
    for header in IDENTITY_HEADERS:
        if header not in headers:
            continue
        index = headers.index(header)
        sftp_value = sftp_row[index] if index < len(sftp_row) else ""
        fabric_value = fabric_row[index] if index < len(fabric_row) else ""
        if sftp_value and sftp_value == fabric_value:
            shared_parts.append(f"{header}={sftp_value}")
            if header not in WEAK_IDENTITY_HEADERS:
                strong_shared_parts.append(f"{header}={sftp_value}")
        elif sftp_value or fabric_value:
            changed_parts.append(f"{header}: {sftp_value} -> {fabric_value}")

    if strong_shared_parts:
        return " | ".join(shared_parts[:3])
    if changed_parts:
        if shared_parts:
            return " | ".join([*shared_parts[:2], *changed_parts[:2]])
        return " | ".join(changed_parts[:3])
    if shared_parts:
        return " | ".join(shared_parts[:3])
    return row_label(headers, sftp_row)


def row_label(headers: list[str], row: list[str]) -> str:
    parts: list[str] = []
    for header in IDENTITY_HEADERS:
        if header not in headers:
            continue
        index = headers.index(header)
        if index < len(row) and row[index]:
            parts.append(f"{header}={row[index]}")
        if len(parts) == 3:
            break
    if parts:
        return " | ".join(parts)
    fallback = [value for value in row[:4] if value]
    return " | ".join(fallback)


def row_signature(row: list[str]) -> str:
    encoded = json.dumps(row, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def detail_row(
    key: ReportKey,
    diff_type: str,
    field: str,
    sftp_value: str,
    fabric_value: str,
    note: str,
    row_key: str = "",
    row_group: str = "",
) -> dict[str, object]:
    return {
        "report_date": f"{key.report_date:%Y-%m-%d}",
        "brand": key.brand,
        "report": key.report,
        "row_key": row_key,
        "row_group": row_group,
        "diff_type": diff_type,
        "field": field,
        "sftp_value": sftp_value,
        "fabric_value": fabric_value,
        "note": note,
    }


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8") as output:
        writer = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_json(path: Path, rows: list[dict[str, object]]) -> None:
    path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")


def write_dashboard(
    path: Path,
    summary_rows: list[dict[str, object]],
    detail_rows: list[dict[str, object]],
    run_context: RunContext,
    artifact_paths: dict[str, Path],
    sftp_source_url: str | None = None,
    fabric_source_url: str | None = None,
) -> None:
    payload = build_dashboard_payload(summary_rows, detail_rows, run_context, artifact_paths, sftp_source_url, fabric_source_url)
    payload_json = json.dumps(payload, ensure_ascii=False).replace("</", "<\\/")
    path.write_text(
        f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>DGE Report Reconciliation</title>
  <style>
    :root {{
      --bg: #f6f7f9;
      --panel: #ffffff;
      --text: #1f2937;
      --muted: #667085;
      --line: #d8dee6;
      --soft-line: #e8edf3;
      --green: #16803c;
      --green-soft: #e8f5ec;
      --red: #bd2b2b;
      --red-soft: #fdecec;
      --amber: #b46100;
      --amber-soft: #fff2dc;
      --blue: #1f5f99;
      --blue-soft: #e7f0f8;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: var(--bg);
      color: var(--text);
      font-family: Arial, sans-serif;
      font-size: 14px;
    }}
    header {{
      background: var(--panel);
      border-bottom: 1px solid var(--line);
      padding: 20px 24px 16px;
      position: sticky;
      top: 0;
      z-index: 5;
    }}
    main {{ padding: 20px 24px 32px; }}
    h1 {{ font-size: 24px; margin: 0 0 6px; letter-spacing: 0; }}
    h2 {{ font-size: 17px; margin: 0 0 12px; }}
    h3 {{ font-size: 14px; margin: 18px 0 8px; }}
    a {{ color: var(--blue); }}
    .muted {{ color: var(--muted); }}
    .topline {{ display: flex; gap: 12px; flex-wrap: wrap; align-items: center; }}
    .run-meta, .artifact-links {{ margin-top: 10px; display: flex; gap: 8px; flex-wrap: wrap; align-items: center; }}
    .meta-chip {{
      border: 1px solid var(--soft-line);
      background: #fbfcfd;
      color: #344054;
      border-radius: 999px;
      padding: 4px 9px;
      font-size: 12px;
      font-weight: 700;
    }}
    .artifact-links a {{
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 5px 8px;
      background: #fff;
      font-size: 12px;
      font-weight: 700;
      text-decoration: none;
    }}
    .source-links {{ margin-top: 10px; display: flex; gap: 12px; flex-wrap: wrap; }}
    .source-links a {{ font-weight: 600; text-decoration: none; }}
    .controls {{ display: grid; gap: 12px; margin-top: 16px; }}
    .control-row {{ display: flex; gap: 8px; flex-wrap: wrap; align-items: center; }}
    .control-label {{ font-size: 12px; font-weight: 700; color: var(--muted); text-transform: uppercase; }}
    .date-filter {{
      display: grid;
      grid-template-columns: minmax(420px, 1fr) auto;
      gap: 12px;
      align-items: center;
    }}
    .date-inputs {{
      display: flex;
      gap: 8px;
      flex-wrap: wrap;
      align-items: center;
    }}
    .date-inputs label {{
      display: grid;
      gap: 4px;
      color: var(--muted);
      font-size: 11px;
      font-weight: 700;
      text-transform: uppercase;
    }}
    input[type="date"] {{
      min-width: 150px;
      border: 1px solid var(--line);
      border-radius: 6px;
      padding: 7px 9px;
      color: var(--text);
      background: #fff;
      font: inherit;
    }}
    .date-summary {{
      justify-self: end;
      color: #344054;
      font-size: 12px;
      font-weight: 700;
      background: #fbfcfd;
      border: 1px solid var(--soft-line);
      border-radius: 999px;
      padding: 6px 10px;
      white-space: nowrap;
    }}
    .matrix-scroll {{
      overflow-x: auto;
      border: 1px solid var(--soft-line);
      border-radius: 8px;
    }}
    .matrix-toolbar {{
      display: flex;
      justify-content: space-between;
      gap: 10px;
      flex-wrap: wrap;
      align-items: center;
      margin-bottom: 10px;
    }}
    .matrix-toolbar .button-group {{
      display: flex;
      gap: 6px;
      flex-wrap: wrap;
      align-items: center;
    }}
    .matrix-table {{ min-width: 720px; }}
    .matrix-table th:first-child, .matrix-table td:first-child {{
      position: sticky;
      left: 0;
      z-index: 1;
      background: #f8fafc;
      min-width: 170px;
      font-weight: 700;
    }}
    .matrix-cell {{
      width: 100%;
      min-width: 104px;
      min-height: 54px;
      display: grid;
      gap: 2px;
      place-items: center;
      border-radius: 6px;
      border: 1px solid transparent;
      font-size: 12px;
      font-weight: 700;
      padding: 6px;
    }}
    .matrix-cell small {{ color: var(--muted); font-weight: 600; }}
    .matrix-cell.status-MATCH {{ background: var(--green-soft); color: var(--green); }}
    .matrix-cell.status-MISMATCH {{ background: var(--red-soft); color: var(--red); }}
    .matrix-cell.status-MISSING_IN_FABRIC,
    .matrix-cell.status-MISSING_IN_SFTP,
    .matrix-cell.status-MISSING_BOTH {{ background: var(--amber-soft); color: var(--amber); }}
    .matrix-empty {{
      color: var(--muted);
      font-size: 12px;
    }}
    .timeline {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(10px, 1fr));
      gap: 3px;
      margin-bottom: 14px;
    }}
    .timeline-day {{
      height: 28px;
      min-width: 10px;
      border: 1px solid var(--soft-line);
      border-radius: 4px;
      background: #eef2f6;
      cursor: pointer;
    }}
    .timeline-day.good {{ background: var(--green-soft); border-color: #9bcfaa; }}
    .timeline-day.issue {{ background: var(--red-soft); border-color: #df9a9a; }}
    .timeline-day.missing {{ background: var(--amber-soft); border-color: #e2bd7d; }}
    .timeline-day.selected {{ outline: 2px solid var(--blue); outline-offset: 1px; }}
    .timeline-legend {{ display: flex; gap: 12px; flex-wrap: wrap; color: var(--muted); font-size: 12px; }}
    .legend-item {{ display: inline-flex; align-items: center; gap: 5px; }}
    .legend-swatch {{ width: 14px; height: 10px; border-radius: 3px; display: inline-block; }}
    .legend-swatch.good {{ background: var(--green-soft); border: 1px solid #9bcfaa; }}
    .legend-swatch.issue {{ background: var(--red-soft); border: 1px solid #df9a9a; }}
    .legend-swatch.missing {{ background: var(--amber-soft); border: 1px solid #e2bd7d; }}
    button {{
      border: 1px solid var(--line);
      background: #fff;
      color: var(--text);
      border-radius: 6px;
      padding: 7px 10px;
      font: inherit;
      cursor: pointer;
    }}
    button.active {{ background: var(--blue-soft); border-color: #7aa7ce; color: #113d63; font-weight: 700; }}
    button.has-issue {{ border-color: #d98a8a; }}
    .grid {{ display: grid; gap: 16px; }}
    .kpis {{ grid-template-columns: repeat(5, minmax(140px, 1fr)); }}
    .panels {{ grid-template-columns: minmax(0, 1.1fr) minmax(0, .9fr); }}
    .section-head {{ display: flex; justify-content: space-between; gap: 12px; align-items: baseline; flex-wrap: wrap; }}
    .section-note {{ color: var(--muted); font-size: 12px; }}
    .panel {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 16px;
      overflow: hidden;
    }}
    .kpi strong {{ display: block; font-size: 28px; line-height: 1.1; margin-top: 4px; }}
    .kpi span {{ font-size: 12px; color: var(--muted); text-transform: uppercase; font-weight: 700; }}
    .kpi.good {{ border-left: 5px solid var(--green); }}
    .kpi.bad {{ border-left: 5px solid var(--red); }}
    .kpi.warn {{ border-left: 5px solid var(--amber); }}
    .status-pill {{
      display: inline-flex;
      align-items: center;
      border-radius: 999px;
      padding: 3px 8px;
      font-size: 12px;
      font-weight: 700;
      line-height: 1.2;
      min-width: max-content;
      white-space: normal;
    }}
    .status-MATCH {{ background: var(--green-soft); color: var(--green); }}
    .status-MISMATCH {{ background: var(--red-soft); color: var(--red); }}
    .status-MISSING_IN_FABRIC, .status-MISSING_IN_SFTP, .status-MISSING_BOTH {{ background: var(--amber-soft); color: var(--amber); }}
    .bar-row, .report-row {{
      display: grid;
      grid-template-columns: 110px minmax(160px, 1fr) 60px;
      gap: 10px;
      align-items: center;
      margin: 10px 0;
    }}
    .bar, .report-bar {{
      height: 18px;
      background: #eef2f6;
      border: 1px solid var(--soft-line);
      border-radius: 4px;
      overflow: hidden;
      display: flex;
    }}
    .bar .ok, .report-bar .ok {{ background: var(--green); }}
    .bar .issue, .report-bar .issue {{ background: var(--red); }}
    .bar .missing, .report-bar .missing {{ background: var(--amber); }}
    .empty-state {{ color: var(--muted); padding: 18px; border: 1px dashed var(--line); border-radius: 8px; background: #fbfcfd; }}
    table {{ border-collapse: collapse; width: 100%; font-size: 13px; background: var(--panel); }}
    th, td {{ border-bottom: 1px solid var(--soft-line); padding: 8px; vertical-align: top; text-align: left; }}
    th {{ background: #f2f5f8; color: #344054; position: sticky; top: 116px; z-index: 2; }}
    tr:hover td {{ background: #fbfcfd; }}
    .summary-table th:first-child, .summary-table td:first-child {{ min-width: 150px; width: 150px; }}
    .summary-table th:nth-child(3), .summary-table td:nth-child(3) {{ min-width: 150px; }}
    .matrix-table th:not(:first-child), .matrix-table td:not(:first-child) {{ text-align: center; }}
    .matrix-table td:last-child {{ text-align: left; }}
    .issue-table td:nth-child(2) {{ font-weight: 700; }}
    code {{ white-space: pre-wrap; word-break: break-word; }}
    .path-cell {{ max-width: 360px; }}
    .actions a {{ display: inline-block; margin-right: 8px; font-weight: 700; text-decoration: none; }}
    .details-table td:nth-child(6), .details-table td:nth-child(7) {{ min-width: 260px; }}
    .group-row td {{
      background: #f8fafc;
      border-top: 2px solid var(--line);
      font-weight: 700;
    }}
    .group-row .muted {{
      display: block;
      font-weight: 400;
      margin-top: 3px;
    }}
    @media (max-width: 980px) {{
      header {{ position: static; }}
      main {{ padding: 14px; }}
      .kpis, .panels {{ grid-template-columns: 1fr; }}
      .date-filter {{ grid-template-columns: 1fr; }}
      .date-summary {{ justify-self: start; white-space: normal; }}
      .bar-row, .report-row {{ grid-template-columns: 1fr; gap: 5px; }}
      th {{ position: static; }}
    }}
  </style>
</head>
<body>
  <header>
    <div class="topline">
      <h1>DGE Report Reconciliation</h1>
      <span id="overallStatus" class="status-pill"></span>
      <span id="generatedAt" class="muted"></span>
    </div>
    <div id="runMeta" class="run-meta"></div>
    <div id="artifactLinks" class="artifact-links"></div>
    <div id="sourceLinks" class="source-links"></div>
    <div class="controls">
      <div class="control-row date-filter" id="dateControls"></div>
      <div class="control-row" id="reportTabs"></div>
    </div>
  </header>

  <main class="grid">
    <section class="grid kpis" id="kpiGrid"></section>
    <section class="panel">
      <div class="section-head">
        <h2>Report x Date Matrix</h2>
        <span class="section-note">Click a cell to filter the dashboard to that report and date.</span>
      </div>
      <div id="dailyMatrix"></div>
    </section>
    <section class="grid panels">
      <div class="panel">
        <div class="section-head">
          <h2>Date Overview</h2>
          <span class="section-note">Whole-period health. Click a day to focus it.</span>
        </div>
        <div id="periodChart"></div>
      </div>
      <div class="panel">
        <div class="section-head">
          <h2>Report Summary</h2>
          <span class="section-note">Aggregated over the selected date range.</span>
        </div>
        <div id="reportChart"></div>
      </div>
    </section>
    <section class="panel">
      <div class="section-head">
        <h2>Summary</h2>
        <span class="section-note">Open file links are available for every checked pair.</span>
      </div>
      <div id="summaryTable"></div>
    </section>
    <section class="panel">
      <div class="section-head">
        <h2>Issue Breakdown</h2>
        <span class="section-note">Use this for data-team follow-up and recurring fixes.</span>
      </div>
      <div id="issueBreakdown"></div>
    </section>
    <section class="panel">
      <div class="section-head">
        <h2>Differences</h2>
        <span class="section-note">Values are listed as SFTP value first, Fabric value second.</span>
      </div>
      <div id="detailsTable"></div>
    </section>
  </main>

  <script id="dashboard-data" type="application/json">{payload_json}</script>
  <script>
    const data = JSON.parse(document.getElementById("dashboard-data").textContent);
    const labels = {{
      game_round_report: "Game round",
      game_summary_report: "Game summary",
      machine_summary_report: "Machine summary",
      pending_transaction: "Pending transaction",
      session_transaction: "Session transaction",
      void_transaction: "Void transaction"
    }};
    const state = {{
      dateMode: "ALL",
      startDate: data.dates[0] || "",
      endDate: data.dates[data.dates.length - 1] || "",
      report: "ALL",
      matrixGranularity: "AUTO"
    }};

    function escapeHtml(value) {{
      return String(value ?? "")
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;")
        .replaceAll("'", "&#39;");
    }}

    function reportLabel(report) {{
      return labels[report] || report;
    }}

    function isIssue(row) {{
      return row.status !== "MATCH";
    }}

    function statusClass(status) {{
      return "status-" + status;
    }}

    function statusLabel(status) {{
      if (status === "MATCH") return "GREEN";
      if (status === "MISMATCH") return "RED";
      return String(status || "").replaceAll("_", " ");
    }}

    function formatDateLabel(value) {{
      if (!value) return "";
      const date = new Date(`${{value}}T00:00:00`);
      if (Number.isNaN(date.getTime())) return value;
      return date.toLocaleDateString(undefined, {{ month: "short", day: "numeric", year: "numeric" }});
    }}

    function shortDateLabel(value) {{
      if (!value) return "";
      const date = new Date(`${{value}}T00:00:00`);
      if (Number.isNaN(date.getTime())) return value;
      return date.toLocaleDateString(undefined, {{ month: "short", day: "numeric" }});
    }}

    function monthLabel(value) {{
      if (!value) return "";
      const date = new Date(`${{value}}-01T00:00:00`);
      if (Number.isNaN(date.getTime())) return value;
      return date.toLocaleDateString(undefined, {{ month: "short", year: "numeric" }});
    }}

    function parseDate(value) {{
      const [year, month, day] = value.split("-").map(Number);
      return new Date(Date.UTC(year, month - 1, day));
    }}

    function isoDate(value) {{
      return value.toISOString().slice(0, 10);
    }}

    function weekKey(value) {{
      const parsed = parseDate(value);
      const day = parsed.getUTCDay() || 7;
      parsed.setUTCDate(parsed.getUTCDate() - day + 1);
      return isoDate(parsed);
    }}

    function selectedDates() {{
      return data.dates.filter(dateInRange);
    }}

    function effectiveMatrixGranularity(dates) {{
      if (state.matrixGranularity !== "AUTO") return state.matrixGranularity;
      if (dates.length <= 45) return "DAY";
      if (dates.length <= 120) return "WEEK";
      return "MONTH";
    }}

    function buildDateBuckets(dates) {{
      const granularity = effectiveMatrixGranularity(dates);
      if (granularity === "DAY") {{
        return dates.map(date => ({{
          key: date,
          label: shortDateLabel(date),
          fullLabel: formatDateLabel(date),
          start: date,
          end: date,
          dates: [date]
        }}));
      }}

      const buckets = new Map();
      dates.forEach(date => {{
        const key = granularity === "MONTH" ? date.slice(0, 7) : weekKey(date);
        if (!buckets.has(key)) buckets.set(key, []);
        buckets.get(key).push(date);
      }});

      return [...buckets.entries()].map(([key, bucketDates]) => {{
        const start = bucketDates[0];
        const end = bucketDates[bucketDates.length - 1];
        return {{
          key,
          label: granularity === "MONTH" ? monthLabel(key) : `${{shortDateLabel(start)}} - ${{shortDateLabel(end)}}`,
          fullLabel: granularity === "MONTH" ? monthLabel(key) : `${{formatDateLabel(start)}} - ${{formatDateLabel(end)}}`,
          start,
          end,
          dates: bucketDates
        }};
      }});
    }}

    function dateInRange(value) {{
      if (state.dateMode === "ALL") return true;
      return value >= state.startDate && value <= state.endDate;
    }}

    function dateRows(date) {{
      return data.summary.filter(row =>
        row.report_date === date &&
        (state.report === "ALL" || row.report === state.report)
      );
    }}

    function setDateRange(mode, startDate, endDate) {{
      const minDate = data.dates[0] || "";
      const maxDate = data.dates[data.dates.length - 1] || "";
      state.dateMode = mode;
      state.startDate = startDate || minDate;
      state.endDate = endDate || maxDate;
      if (state.startDate > state.endDate) {{
        const swap = state.startDate;
        state.startDate = state.endDate;
        state.endDate = swap;
      }}
      renderAll();
    }}

    function sourceLink(label, url) {{
      if (!url) return "";
      return `<a href="${{escapeHtml(url)}}">${{escapeHtml(label)}}</a>`;
    }}

    function artifactLink(label, url) {{
      if (!url) return "";
      return `<a href="${{escapeHtml(url)}}">${{escapeHtml(label)}}</a>`;
    }}

    function matchingRows() {{
      return data.summary.filter(row =>
        dateInRange(row.report_date) &&
        (state.report === "ALL" || row.report === state.report)
      );
    }}

    function matchingDetails() {{
      return data.details.filter(row =>
        dateInRange(row.report_date) &&
        (state.report === "ALL" || row.report === state.report)
      );
    }}

    function groupRows(rows, key) {{
      return rows.reduce((acc, row) => {{
        const value = row[key];
        if (!acc[value]) acc[value] = [];
        acc[value].push(row);
        return acc;
      }}, {{}});
    }}

    function renderHeader() {{
      document.getElementById("generatedAt").textContent = `Generated ${{data.generatedAt}}`;
      const overallIssue = data.summary.some(isIssue);
      const status = overallIssue ? "RED" : "GREEN";
      const statusEl = document.getElementById("overallStatus");
      statusEl.textContent = status;
      statusEl.className = "status-pill " + (overallIssue ? "status-MISMATCH" : "status-MATCH");
      document.getElementById("runMeta").innerHTML = [
        ["Mode", data.run.mode],
        ["Run", data.run.id],
        ["Business dates", data.run.businessDateRange],
        ["Brand", data.run.brand],
        ["Reports", data.run.reports],
        ["Rows", data.run.rowOrderPolicy]
      ].map(([label, value]) => `<span class="meta-chip">${{escapeHtml(label)}}: ${{escapeHtml(value)}}</span>`).join("");
      document.getElementById("artifactLinks").innerHTML = Object.entries(data.artifacts || {{}})
        .map(([label, url]) => artifactLink(label, url))
        .filter(Boolean)
        .join("");
      document.getElementById("sourceLinks").innerHTML = [
        sourceLink("SFTP source folder", data.sources.sftp),
        sourceLink("Fabric source folder", data.sources.fabric)
      ].filter(Boolean).join(" | ");
    }}

    function renderTabs() {{
      const dateContainer = document.getElementById("dateControls");
      const reportContainer = document.getElementById("reportTabs");
      renderDateControls(dateContainer);
      reportContainer.innerHTML = `<span class="control-label">Report</span>` +
        tabButton("report", "ALL", "All reports") +
        data.reports.map(report => tabButton("report", report, reportLabel(report))).join("");
    }}

    function renderDateControls(container) {{
      const minDate = data.dates[0] || "";
      const maxDate = data.dates[data.dates.length - 1] || "";
      const selectedCount = data.dates.filter(dateInRange).length;
      const label = state.dateMode === "ALL"
        ? `Full period: ${{formatDateLabel(minDate)}} - ${{formatDateLabel(maxDate)}} (${{selectedCount}} dates)`
        : `${{formatDateLabel(state.startDate)}} - ${{formatDateLabel(state.endDate)}} (${{selectedCount}} date${{selectedCount === 1 ? "" : "s"}})`;

      container.innerHTML = `
        <div class="date-inputs">
          <span class="control-label">Date range</span>
          <button data-date-all="true" class="${{state.dateMode === "ALL" ? "active" : ""}}">Full period</button>
          <label>From <input id="dateFrom" type="date" min="${{escapeHtml(minDate)}}" max="${{escapeHtml(maxDate)}}" value="${{escapeHtml(state.startDate || minDate)}}"></label>
          <label>To <input id="dateTo" type="date" min="${{escapeHtml(minDate)}}" max="${{escapeHtml(maxDate)}}" value="${{escapeHtml(state.endDate || maxDate)}}"></label>
          <button data-date-apply="range" class="${{state.dateMode === "CUSTOM" ? "active" : ""}}">Apply</button>
        </div>
        <div class="date-summary">${{escapeHtml(label)}}</div>
      `;
    }}

    function tabButton(kind, value, label) {{
      const rows = data.summary.filter(row =>
        value === "ALL" || row.report === value
      );
      const issueClass = rows.some(isIssue) ? " has-issue" : "";
      const activeClass = state[kind] === value ? " active" : "";
      return `<button class="${{activeClass}}${{issueClass}}" data-kind="${{kind}}" data-value="${{escapeHtml(value)}}">${{escapeHtml(label)}}</button>`;
    }}

    function renderKpis() {{
      const rows = matchingRows();
      const total = rows.length;
      const matched = rows.filter(row => row.status === "MATCH").length;
      const mismatched = rows.filter(row => row.status === "MISMATCH").length;
      const missing = rows.filter(row => row.status.startsWith("MISSING")).length;
      const diffs = rows.reduce((sum, row) => sum + Number(row.difference_count || 0), 0);
      const cards = [
        ["Total checks", total, ""],
        ["Green", matched, "good"],
        ["Red", mismatched, "bad"],
        ["Missing", missing, "warn"],
        ["Diffs", diffs, diffs ? "bad" : "good"]
      ];
      document.getElementById("kpiGrid").innerHTML = cards.map(([label, value, klass]) =>
        `<div class="panel kpi ${{klass}}"><span>${{escapeHtml(label)}}</span><strong>${{escapeHtml(value)}}</strong></div>`
      ).join("");
    }}

    function renderDailyMatrix() {{
      const container = document.getElementById("dailyMatrix");
      const dates = selectedDates();
      const reports = state.report === "ALL" ? data.reports : [state.report];
      if (!dates.length || !reports.length) {{
        container.innerHTML = `<div class="empty-state">No dates in the selected range.</div>`;
        return;
      }}
      const buckets = buildDateBuckets(dates);
      const granularity = effectiveMatrixGranularity(dates);
      const header = buckets.map(bucket => `<th title="${{escapeHtml(bucket.fullLabel)}}">${{escapeHtml(bucket.label)}}</th>`).join("");
      const body = reports.map(report => {{
        const cells = buckets.map(bucket => {{
          const rows = bucket.dates
            .map(date => data.summary.find(item => item.report === report && item.report_date === date))
            .filter(Boolean);
          if (!rows.length) return `<td><span class="matrix-empty">No check</span></td>`;
          return matrixCell(report, bucket, rows, granularity);
        }}).join("");
        return `<tr><td>${{escapeHtml(reportLabel(report))}}</td>${{cells}}</tr>`;
      }}).join("");
      container.innerHTML = `
        <div class="matrix-toolbar">
          <div class="button-group">
            <span class="control-label">Group by</span>
            ${{["AUTO", "DAY", "WEEK", "MONTH"].map(value => `<button data-matrix-granularity="${{value}}" class="${{state.matrixGranularity === value ? "active" : ""}}">${{escapeHtml(value === "AUTO" ? `Auto (${{granularity.toLowerCase()}})` : value[0] + value.slice(1).toLowerCase())}}</button>`).join("")}}
          </div>
          <span class="section-note">${{escapeHtml(buckets.length)}} column${{buckets.length === 1 ? "" : "s"}} over ${{dates.length}} selected date${{dates.length === 1 ? "" : "s"}}</span>
        </div>
        <div class="matrix-scroll"><table class="matrix-table">
        <thead><tr><th>Report</th>${{header}}</tr></thead>
        <tbody>${{body}}</tbody>
      </table></div>`;
    }}

    function matrixCell(report, bucket, rows, granularity) {{
      const ok = rows.filter(row => row.status === "MATCH").length;
      const mismatch = rows.filter(row => row.status === "MISMATCH").length;
      const missing = rows.filter(row => row.status.startsWith("MISSING")).length;
      const diffs = rows.reduce((sum, row) => sum + Number(row.difference_count || 0), 0);
      const status = missing ? "MISSING_IN_FABRIC" : mismatch ? "MISMATCH" : "MATCH";
      const top = granularity === "DAY" ? statusLabel(rows[0].status) : `${{ok}}/${{rows.length}} green`;
      const bottom = granularity === "DAY" ? `${{diffs}} diffs` : `${{mismatch + missing}} issue day${{mismatch + missing === 1 ? "" : "s"}}, ${{diffs}} diffs`;
      return `<td><button class="matrix-cell ${{statusClass(status)}}" data-date-start="${{escapeHtml(bucket.start)}}" data-date-end="${{escapeHtml(bucket.end)}}" data-report-value="${{escapeHtml(report)}}" title="${{escapeHtml(bucket.fullLabel)}}: ${{escapeHtml(bottom)}}">
        <span>${{escapeHtml(top)}}</span>
        <small>${{escapeHtml(bottom)}}</small>
      </button></td>`;
    }}

    function renderPeriodChart() {{
      const container = document.getElementById("periodChart");
      const rowsByDate = groupRows(data.summary.filter(row => state.report === "ALL" || row.report === state.report), "report_date");
      const timeline = `<div class="timeline" aria-label="Date health timeline">${{data.dates.map(date => {{
        const rows = rowsByDate[date] || [];
        const klass = rows.some(row => row.status.startsWith("MISSING")) ? "missing" : rows.some(row => row.status === "MISMATCH") ? "issue" : "good";
        const selected = dateInRange(date) ? " selected" : "";
        const title = `${{date}}: ${{rows.filter(row => row.status === "MATCH").length}} green, ${{rows.filter(row => row.status === "MISMATCH").length}} mismatch, ${{rows.filter(row => row.status.startsWith("MISSING")).length}} missing`;
        return `<button class="timeline-day ${{klass}}${{selected}}" data-date-single="${{escapeHtml(date)}}" title="${{escapeHtml(title)}}" aria-label="${{escapeHtml(title)}}"></button>`;
      }}).join("")}}</div>`;
      const selectedCount = data.dates.filter(dateInRange).length;
      const legend = `<div class="timeline-legend">
        <span class="legend-item"><span class="legend-swatch good"></span> Green</span>
        <span class="legend-item"><span class="legend-swatch issue"></span> Mismatch</span>
        <span class="legend-item"><span class="legend-swatch missing"></span> Missing</span>
        <span>${{selectedCount}} of ${{data.dates.length}} dates selected</span>
      </div>`;
      container.innerHTML = data.dates.length ? timeline + legend : `<div class="empty-state">No period data.</div>`;
    }}

    function renderReportChart() {{
      const container = document.getElementById("reportChart");
      const rowsByReport = groupRows(data.summary.filter(row => dateInRange(row.report_date)), "report");
      const html = data.reports.map(report => {{
        const rows = rowsByReport[report] || [];
        const total = rows.length || 1;
        const ok = rows.filter(row => row.status === "MATCH").length;
        const mismatch = rows.filter(row => row.status === "MISMATCH").length;
        const missing = rows.filter(row => row.status.startsWith("MISSING")).length;
        return `<div class="report-row">
          <button data-kind="report" data-value="${{escapeHtml(report)}}" class="${{state.report === report ? "active" : ""}}">${{escapeHtml(reportLabel(report))}}</button>
          <div class="report-bar" title="${{ok}} green, ${{mismatch}} mismatch, ${{missing}} missing">
            <div class="ok" style="width:${{ok / total * 100}}%"></div>
            <div class="issue" style="width:${{mismatch / total * 100}}%"></div>
            <div class="missing" style="width:${{missing / total * 100}}%"></div>
          </div>
          <span>${{ok}}/${{rows.length}}</span>
        </div>`;
      }}).join("");
      container.innerHTML = html || `<div class="empty-state">No report data.</div>`;
    }}

    function renderSummaryTable() {{
      const rows = matchingRows();
      if (!rows.length) {{
        document.getElementById("summaryTable").innerHTML = `<div class="empty-state">No checks for the selected filters.</div>`;
        return;
      }}
      document.getElementById("summaryTable").innerHTML = `<table class="summary-table">
        <thead>
          <tr><th>Status</th><th>Date</th><th>Report</th><th>SFTP rows</th><th>Fabric rows</th><th>Diffs</th><th>Comments</th><th>Files</th></tr>
        </thead>
        <tbody>${{rows.map(summaryRow).join("")}}</tbody>
      </table>`;
    }}

    function summaryRow(row) {{
      const sftp = row.sftp_file_url ? `<a href="${{escapeHtml(row.sftp_file_url)}}">SFTP</a>` : "";
      const fabric = row.fabric_file_url ? `<a href="${{escapeHtml(row.fabric_file_url)}}">Fabric</a>` : "";
      return `<tr>
        <td><span class="status-pill ${{statusClass(row.status)}}">${{escapeHtml(statusLabel(row.status))}}</span></td>
        <td>${{escapeHtml(row.report_date)}}</td>
        <td>${{escapeHtml(reportLabel(row.report))}}</td>
        <td>${{escapeHtml(row.sftp_rows)}}</td>
        <td>${{escapeHtml(row.fabric_rows)}}</td>
        <td>${{escapeHtml(row.difference_count)}}</td>
        <td>${{escapeHtml(row.comments)}}</td>
        <td class="actions">${{sftp}} ${{fabric}}<br><code>${{escapeHtml(row.sftp_path)}}<br>${{escapeHtml(row.fabric_path)}}</code></td>
      </tr>`;
    }}

    function renderIssueBreakdown() {{
      const container = document.getElementById("issueBreakdown");
      const details = matchingDetails();
      if (!details.length) {{
        container.innerHTML = `<div class="empty-state">No differences for the selected range.</div>`;
        return;
      }}
      const counts = new Map();
      details.forEach(row => {{
        const key = [reportLabel(row.report), row.diff_type, row.field || row.diff_type].join("|");
        if (!counts.has(key)) {{
          counts.set(key, {{
            report: reportLabel(row.report),
            type: row.diff_type,
            field: row.field || row.diff_type,
            count: 0
          }});
        }}
        counts.get(key).count += 1;
      }});
      const rows = [...counts.values()].sort((a, b) => b.count - a.count).slice(0, 20);
      container.innerHTML = `<table class="issue-table">
        <thead><tr><th>Report</th><th>Type</th><th>Field</th><th>Count</th></tr></thead>
        <tbody>${{rows.map(row => `<tr><td>${{escapeHtml(row.report)}}</td><td>${{escapeHtml(row.type)}}</td><td>${{escapeHtml(row.field)}}</td><td>${{escapeHtml(row.count)}}</td></tr>`).join("")}}</tbody>
      </table>`;
    }}

    function summaryFor(row) {{
      return data.summary.find(item =>
        item.report_date === row.report_date &&
        item.report === row.report &&
        item.brand === row.brand
      ) || null;
    }}

    function fileActionsFor(row) {{
      const summary = summaryFor(row);
      if (!summary) return "";
      const sftp = summary.sftp_file_url ? `<a href="${{escapeHtml(summary.sftp_file_url)}}">Open SFTP file</a>` : "";
      const fabric = summary.fabric_file_url ? `<a href="${{escapeHtml(summary.fabric_file_url)}}">Open Fabric file</a>` : "";
      return [sftp, fabric].filter(Boolean).join(" ");
    }}

    function renderDetailsTable() {{
      const rows = matchingDetails();
      if (!rows.length) {{
        document.getElementById("detailsTable").innerHTML = `<div class="empty-state">No differences for the selected filters.</div>`;
        return;
      }}
      const workbookRows = rows.filter(row => ["metadata", "file", "row_count"].includes(row.diff_type));
      const valueRows = rows.filter(row => !["metadata", "file", "row_count"].includes(row.diff_type));
      const workbookTable = workbookRows.length ? `<h3>Workbook / schema checks</h3>
        <table class="details-table">
          <thead>
            <tr><th>Date</th><th>Report</th><th>Type</th><th>Field</th><th>Note</th><th>SFTP value</th><th>Fabric value</th><th>Files</th></tr>
          </thead>
          <tbody>${{workbookRows.slice(0, 200).map(workbookDetailRow).join("")}}</tbody>
        </table>` : "";
      const valueTable = valueRows.length ? `<h3>Row / value checks</h3>
        <table class="details-table">
        <thead>
          <tr><th>Date</th><th>Report</th><th>Row</th><th>Type</th><th>Field</th><th>Note</th><th>SFTP value</th><th>Fabric value</th></tr>
        </thead>
          <tbody>${{renderGroupedValueRows(valueRows.slice(0, 500))}}</tbody>
        </table>` : "";
      document.getElementById("detailsTable").innerHTML = workbookTable + valueTable;
    }}

    function workbookDetailRow(row) {{
      return `<tr>
        <td>${{escapeHtml(row.report_date)}}</td>
        <td>${{escapeHtml(reportLabel(row.report))}}</td>
        <td>${{escapeHtml(row.diff_type)}}</td>
        <td>${{escapeHtml(row.field)}}</td>
        <td>${{escapeHtml(row.note)}}</td>
        <td><code>${{escapeHtml(row.sftp_value)}}</code></td>
        <td><code>${{escapeHtml(row.fabric_value)}}</code></td>
        <td class="actions">${{fileActionsFor(row)}}</td>
      </tr>`;
    }}

    function renderGroupedValueRows(rows) {{
      const groups = [];
      const indexByKey = new Map();
      rows.forEach(row => {{
        const groupKey = row.row_group || [row.report_date, row.report, row.row_key || row.field, row.diff_type === "cell" ? "cell" : row.field].join("|");
        if (!indexByKey.has(groupKey)) {{
          indexByKey.set(groupKey, groups.length);
          groups.push({{
            date: row.report_date,
            report: row.report,
            rowKey: row.row_key || row.field,
            kind: row.diff_type === "cell" ? "cell" : row.field,
            rows: []
          }});
        }}
        groups[indexByKey.get(groupKey)].rows.push(row);
      }});

      return groups.map(group => {{
        const changedFields = group.rows.filter(row => row.diff_type === "cell").length;
        const groupNote = changedFields > 1
          ? `${{changedFields}} changed fields in this same row`
          : changedFields === 1
            ? "1 changed field in this row"
            : group.rows[0]?.note || group.kind;
        const actions = fileActionsFor(group.rows[0]);
        const actionLine = actions ? `<span class="muted actions">${{actions}}</span>` : "";
        const header = `<tr class="group-row"><td>${{escapeHtml(group.date)}}</td><td>${{escapeHtml(reportLabel(group.report))}}</td><td colspan="6">${{escapeHtml(group.rowKey)}}<span class="muted">${{escapeHtml(groupNote)}}</span>${{actionLine}}</td></tr>`;
        return header + group.rows.map(valueDetailRow).join("");
      }}).join("");
    }}

    function valueDetailRow(row) {{
      return `<tr>
        <td></td>
        <td></td>
        <td></td>
        <td>${{escapeHtml(row.diff_type)}}</td>
        <td>${{escapeHtml(row.field)}}</td>
        <td>${{escapeHtml(row.note)}}</td>
        <td><code>${{escapeHtml(row.sftp_value)}}</code></td>
        <td><code>${{escapeHtml(row.fabric_value)}}</code></td>
      </tr>`;
    }}

    function renderAll() {{
      renderHeader();
      renderTabs();
      renderKpis();
      renderDailyMatrix();
      renderPeriodChart();
      renderReportChart();
      renderSummaryTable();
      renderIssueBreakdown();
      renderDetailsTable();
    }}

    document.addEventListener("click", event => {{
      const allDateButton = event.target.closest("button[data-date-all]");
      if (allDateButton) {{
        setDateRange("ALL", data.dates[0] || "", data.dates[data.dates.length - 1] || "");
        return;
      }}

      const applyButton = event.target.closest("button[data-date-apply]");
      if (applyButton) {{
        const from = document.getElementById("dateFrom")?.value || state.startDate;
        const to = document.getElementById("dateTo")?.value || state.endDate;
        setDateRange("CUSTOM", from, to);
        return;
      }}

      const granularityButton = event.target.closest("button[data-matrix-granularity]");
      if (granularityButton) {{
        state.matrixGranularity = granularityButton.dataset.matrixGranularity;
        renderAll();
        return;
      }}

      const rangeButton = event.target.closest("button[data-date-start][data-date-end]");
      if (rangeButton) {{
        if (rangeButton.dataset.reportValue) {{
          state.report = rangeButton.dataset.reportValue;
        }}
        setDateRange("CUSTOM", rangeButton.dataset.dateStart, rangeButton.dataset.dateEnd);
        return;
      }}

      const singleDateButton = event.target.closest("button[data-date-single]");
      if (singleDateButton) {{
        const date = singleDateButton.dataset.dateSingle;
        if (singleDateButton.dataset.reportValue) {{
          state.report = singleDateButton.dataset.reportValue;
        }}
        setDateRange("CUSTOM", date, date);
        return;
      }}

      const button = event.target.closest("button[data-kind]");
      if (!button) return;
      state[button.dataset.kind] = button.dataset.value;
      renderAll();
    }});

    renderAll();
  </script>
</body>
</html>
""",
        encoding="utf-8",
    )


def build_dashboard_payload(
    summary_rows: list[dict[str, object]],
    detail_rows: list[dict[str, object]],
    run_context: RunContext,
    artifact_paths: dict[str, Path],
    sftp_source_url: str | None,
    fabric_source_url: str | None,
) -> dict[str, object]:
    enriched_summary = []
    for row in summary_rows:
        enriched_row = dict(row)
        enriched_row["sftp_file_url"] = file_href(str(row["sftp_path"]))
        enriched_row["fabric_file_url"] = file_href(str(row["fabric_path"]))
        enriched_summary.append(enriched_row)

    return {
        "generatedAt": run_context.generated_at,
        "run": {
            "id": run_context.run_id,
            "mode": run_context.mode,
            "businessDateRange": run_context.business_date_range,
            "brand": run_context.brand_label,
            "reports": run_context.report_label,
            "rowOrderPolicy": run_context.row_order_policy,
            "fileSuffix": run_context.file_suffix,
        },
        "artifacts": {
            "Dashboard for this run": file_href(str(artifact_paths["dashboard_archived"])),
            "Latest dashboard": file_href(str(artifact_paths["dashboard_latest"])),
            "Summary CSV": file_href(str(artifact_paths["summary_archived"])),
            "Details CSV": file_href(str(artifact_paths["details_archived"])),
            "Summary JSON": file_href(str(artifact_paths["summary_json_archived"])),
            "Details JSON": file_href(str(artifact_paths["details_json_archived"])),
        },
        "sources": {
            "sftp": sftp_source_url or "",
            "fabric": fabric_source_url or "",
        },
        "dates": sorted({str(row["report_date"]) for row in summary_rows}),
        "reports": sorted({str(row["report"]) for row in summary_rows}),
        "summary": enriched_summary,
        "details": detail_rows,
    }


def file_href(path_value: str) -> str:
    if not path_value:
        return ""
    if not path_value.lower().startswith(("http://", "https://", "file://")):
        try:
            return Path(path_value).resolve().as_uri()
        except ValueError:
            return path_value
    return path_value


def utc_now_iso() -> str:
    return datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z")


if __name__ == "__main__":
    raise SystemExit(main())
